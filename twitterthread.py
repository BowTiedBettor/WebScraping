import requests
from datetime import datetime, timedelta
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from traceback import print_exc
import json
import pandas as pd
from openpyxl import load_workbook

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


class Unibet:
    """
    Horse racing scraper,
    - collect odds for SWE & FRA horse racing
    - write to Excel
    - await odds releases -> send notifications
    """
    def __init__(self):
        self.headers = {
                        'User-Agent': 'ADD SOME USER AGENT HERE',
                        'Accept': '*/*',
                        'Accept-Language': 'en-GB,en;q=0.5',
                        'Referer': 'https://www.unibet.se/',
                        'content-type': 'application/json',
                        'Origin': 'https://www.unibet.se',
                        'Connection': 'keep-alive',
                        'Sec-Fetch-Dest': 'empty',
                        'Sec-Fetch-Mode': 'cors',
                        'Sec-Fetch-Site': 'cross-site',
                    }

    def get_meeting(self, track: str, races: list, countrycode = "SWE"):
        """
        Locates and returns the event_ids for the races in the given meeting
        """
        # apparently a maximum of 4 days diff or the server won't respond correctly
        start = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        until = (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d')

        try:
            response = requests.get(f"https://rsa.unibet.co.uk/api/v1/graphql?operationName=MeetingsByDateRange&variables=%7B%22startDateTime%22%3A%22{start}T23%3A00%3A00.000Z%22%2C%22endDateTime%22%3A%22{until}T23%3A00%3A00.000Z%22%2C%22countryCodes%22%3A%22{countrycode}%22%2C%22raceTypes%22%3A%5B%22H%22%5D%7D&extensions=%7B%22persistedQuery%22%3A%7B%22version%22%3A1%2C%22sha256Hash%22%3A%22b975a015a4d4e4dc298ebd6dd43e988ebf03fb940e2bb719478b59bde7bbffd6%22%7D%7D", headers=self.headers, verify = False)
            resp_json = response.json()

            meetings = resp_json['data']['viewer']['meetingsByDateRange']

            for meet in meetings:
              if meet['name'] == track:
                events = meet['events']
                break

            eventkeys = []
            for event in events:
              race = int(event['eventKey'].split(".")[-1])
              if race >= races[0] and race <= races[1]:
                eventkeys.append(event['eventKey'])
        except:
            print_exc()
            return None

        return eventkeys

    def scrape_wp(self, track: str, races: list, countrycode = "SWE"):
        """
        Scrapes the win and place markets
        :param races list: is expected to be on the form [start_race, end_race]
        """
        eventkeys = self.get_meeting(track = track, races = races, countrycode = countrycode)

        if eventkeys:
            # creates an empty list that will hold the pandas dataframes [1 per race]
            pd_list = []

            for eventkey in eventkeys:
                # Creates an empty dataframe for the race
                race_df = pd.DataFrame(
                    columns=["PostPos", "Horse", "Win", "Place"])
                try:
                    response = requests.get(f'https://rsa.unibet.co.uk/api/v1/graphql?operationName=EventQuery&variables=%7B%22clientCountryCode%22%3A%22SE%22%2C%22eventKey%22%3A%22{eventkey}%22%2C%22fetchTRC%22%3Afalse%7D&extensions=%7B%22persistedQuery%22%3A%7B%22version%22%3A1%2C%22sha256Hash%22%3A%228fee3aa36e812c03f4cb039cb2f2c2defc919ea7a89b9b942e5e79588260a8b7%22%7D%7D',
                      headers=self.headers, verify = False)
                    competitors = response.json()['data']['viewer']['event']['competitors']

                    with open('comps.json', 'w') as file:
                        json.dump(competitors, file)

                    for comp in competitors:
                        post_pos = comp['startPos']
                        name = comp['name']
                        prices = comp['prices']
                        if prices:
                            # fix this tmrw when odds have been released such that it collects win and place odds
                            pass
                        else:
                            win_odds = 0.00
                            place_odds = 0.00

                        dummy_df = race_df
                        new_row = pd.DataFrame(
                            [[post_pos, name, win_odds, place_odds]], columns=race_df.columns)
                        race_df = pd.concat(
                            [dummy_df, new_row], ignore_index=True)
                except:
                    print_exc()
                    pass

                # when the race has been completed, sort on post position and append to pd_list
                sorted_df = race_df.sort_values(by="PostPos")
                pd_list.append(sorted_df)

            return pd_list

    def to_excel(self, pd_list: list, file: str, sheet_name: str):
        """
        Dumps the list of pandas frames into an excel file
        """
        current_col = pd.read_excel(file, sheet_name = sheet_name).shape[1]
        time = datetime.now().strftime('%H:%M:%S')

        # insert the time the odds was pulled in the correct cell
        workbook = load_workbook(file)
        worksheet = workbook[sheet_name]
        if current_col == 0:
            worksheet.cell(row=1, column=2).value = time
        else:
            worksheet.cell(row=1, column = current_col + 4).value = time
        workbook.save(file)

        # insert the scraped odds
        i = 0
        for racedf in pd_list:
            with pd.ExcelWriter(file, mode='a', if_sheet_exists='overlay', engine = 'openpyxl') as writer:
                if current_col == 0:
                    if "V&P" in sheet_name:
                        racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 18 * i, startcol=0, index = False)
                        i += 1
                    elif "H2H" in sheet_name:
                        racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 5 * i, startcol=0, index = False)
                        i += 1
                else:
                    if "V&P" in sheet_name:
                        racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 18 * i, startcol=current_col + 2, index = False)
                        i += 1
                    elif "H2H" in sheet_name:
                        racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 5 * i, startcol=current_col + 2, index = False)
                        i += 1


u = Unibet()
scraped_races = u.scrape_wp(track = "Solvalla", races = [5,9])
u.to_excel(pd_list = scraped_races, file = "FILE PATH HERE", sheet_name = "Unibet - V&P")

